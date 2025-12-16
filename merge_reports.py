# !!! make sure u close all the excel files before running this !!!

import pandas as pd
import win32com.client as win32
import glob
import os
import config
import private
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

folder_path = private.PATH
output_file = 'MasterReport.xlsx'

def import_original_sheets(master_path, source_folder):
    if not win32:
        return
    print(f"Importing original sheets")
    
    master_path = os.path.abspath(master_path)
    source_folder = os.path.abspath(source_folder)
    if not os.path.exists(master_path):
        print(f"[ERROR] Master file not found at {master_path}")
        return
    
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = True
    excel.DisplayAlerts = False
    
    try:
        wb_master = excel.Workbooks.Open(master_path)
        
        # grab all the excel files
        all_files = glob.glob(os.path.join(source_folder, "*.xlsx"))
        
        for filename in all_files:
            # skip master file
            if os.path.abspath(filename) == master_path:
                continue
            # skip excel temp lock files
            if os.path.basename(filename).startswith('~$'):
                continue
            
            try:
                wb_source = excel.Workbooks.Open(filename)
                
                # copy the excel sheets to master
                for i, sheet in enumerate(wb_source.Sheets):
                    original_sheet_count = wb_master.Sheets.Count
                    sheet.Copy(After=wb_master.Sheets(original_sheet_count))
                    
                    if wb_master.Sheets.Count > original_sheet_count:
                        try:
                            # remove extension
                            base_name = os.path.splitext(os.path.basename(filename))[0]
                            # if theres multiple sheets, we'll append the index to avoid name collision
                            if wb_source.Sheets.Count > 1:
                                new_name = f"{base_name[:28]}_{i+1}"
                            else:
                                new_name = base_name[:31] # excel name limit is 31
                                
                            wb_master.Sheets(wb_master.Sheets.Count).Name = new_name
                            print(f"[SUCCESS] Finished sheet importing function from {os.path.basename(filename)}")
                        except Exception as error:
                            print(f"[ERROR] Could not rename sheet: {error}")
                    else:
                        print(f"[ERROR] Sheet copy failed for {os.path.basename(filename)}")  
                wb_source.Close(SaveChanges=False)
            except Exception as error:
                print(f"[ERROR] Error processing {filename}: {error}")
        
        wb_master.Save()
        wb_master.Close()
    except Exception as error:
        print(f"[ERROR] Could not import excel sheets: {error}")
    finally:
        try:
            excel.Quit()
        except:
            pass

def merge_excel_sheets(path, output_name):
    all_files = glob.glob(os.path.join(path, "*.xlsx")) # os.path.join to account for Mac OS
    print(f"{len(all_files)} excel files to process")
    all_data = [] # hold tables from each file
    
    for filename in all_files:
        if os.path.basename(filename) == output_name:
            continue
        
        try:
            wb = load_workbook(filename, data_only=True) # data_only=True ignores formulas
            sheet = wb.active
            row_data = {}
            row_data['Source File'] = os.path.basename(filename)
            
            # extract cells
            for col_name, cell_address in config.extraction_cells.items():
                try:
                    # check for list
                    if isinstance(cell_address, list):
                        values = []
                        for addr in cell_address:
                            val = sheet[addr].value
                            if val:
                                values.append(str(val))
                        row_data[col_name] = " ".join(values) if values else None
                    else:
                        val = sheet[cell_address].value
                        row_data[col_name] = val
                except Exception as error:
                    row_data[col_name] = None
                    print(f"[ERROR] Could not read cell {cell_address} in {filename}")
            all_data.append(row_data)
            print(f"Read {os.path.basename(filename)}")
            wb.close()
        except Exception as error:
            row_data[col_name] = None
            print(f"[ERROR] Could not read {filename}: {error}")
        
    # combine tables into one
    if all_data:
        combined_df = pd.DataFrame(all_data)
        
        # export to new excel file
        output_path = os.path.join(path, output_name)
        
        # format output
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            combined_df.to_excel(writer, index=False, sheet_name='Master Report')
            worksheet = writer.sheets['Master Report']
            
            # auto calc column width
            for i, column in enumerate(combined_df.columns):
                max_len = len(str(column))
                if not combined_df.empty:
                    data_len = combined_df[column].astype(str).map(len).max()
                    if not pd.isna(data_len):
                        max_len = max(max_len, data_len)
                
                column_letter = get_column_letter(i + 1)
                
                # formatting requirements
                is_source_file = column == "Source File"
                should_center = column in config.centered_columns
                should_wrap = max_len + 2 > 50
                should_green = column in getattr(config, 'green_columns', [])
                should_red = column in getattr(config, 'red_columns', []) # btw this only turns red if there is text present in the cell (some days theres no event)
                
                # column width wrapping
                if should_wrap:
                    worksheet.column_dimensions[column_letter].width = 50
                else:
                    worksheet.column_dimensions[column_letter].width = max_len + 2
                    
                # green highlighting (for income col)
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                # red highlighting (for daily events)
                red_fill = PatternFill(start_color="FFC1C1", end_color="FFC1C1", fill_type="solid")
                    
                # apply
                for cell in worksheet[column_letter]:
                    # default
                    horiz = None
                    vert = None
                    
                    
                    if should_center:
                        horiz = 'center'
                        vert = 'center'
                    elif is_source_file:
                        vert = 'center'
                        
                    if should_wrap:
                        if not vert: vert = 'center'
                        
                    if should_center or should_wrap or is_source_file:
                        cell.alignment = Alignment(horizontal=horiz, vertical=vert, wrap_text=should_wrap)
                
                    if should_green:
                        cell.fill = green_fill
                        
                    if should_red and cell.value: # cell.value to check for text inside the cell
                        cell.fill = red_fill
                
        print(f"[SUCCESS] All files merged into: {output_path}")
    else:
        print(f"[ERROR] No data found to merge")

if __name__ == "__main__":
    merge_excel_sheets(folder_path, output_file)
    import_original_sheets(os.path.join(folder_path, output_file), folder_path)